import socket
import sys
from datetime import datetime, time

import cv2
import psycopg2
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from PySide6.QtCore import QDate, QRegularExpression, Qt, QThread, QTimer, Signal
from PySide6.QtGui import QColor, QImage, QPixmap, QRegularExpressionValidator
from PySide6.QtWidgets import (
    QApplication,
    QComboBox,
    QDateEdit,
    QDialog,
    QFileDialog,
    QGridLayout,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QPushButton,
    QSizePolicy,
    QStackedWidget,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)
#===================Helper Function============
def clean_text(s):
    if s is None:
        return ""
    return s.replace("\x00", "").strip()

# ================= DB CONFIG =================
DB = dict(
    dbname="camera_inspection",
    user="postgres",
    password="1234",
    host="localhost",
    port=5432,
)
TABLE = "camera_inspection"
last_data=None
# ================= SOCKET THREAD (ADDED) =================
class FHVSocketThread(QThread):
    data_received = Signal(str)
    def __init__(self):
        super().__init__()          #  REQUIRED
        self.running = True         #  REQUIRED
        self.waiting_for_user = False



    def run(self):
        HOST = "172.21.2.11"
        PORT = 9876

        while self.running:
            try:
                print("SOCKET LISTENING")
                server = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                server.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
                server.bind((HOST, PORT))
                server.listen(1)

                conn, _ = server.accept()
                print("SOCKET CONNECTED")

                while self.running:
                    if self.waiting_for_user:
                        self.msleep(100)
                        continue

                    conn.sendall(b"M\r\n")
                    data = conn.recv(4096)

                    if not data:
                        raise ConnectionResetError("Camera closed connection")

                    msg = data.decode("ascii", errors="ignore")
                    msg = msg.replace("\x00", "").strip()

                    if msg in ("ER", "OK", "0"):
                        continue

                    if len(msg) != 23 or not msg.isalnum():
                        continue

                    print("VALID SOCKET:", msg)

                    self.waiting_for_user = True
                    self.msleep(3000)  # 3s delay
                    self.data_received.emit(msg)

            except Exception as e:
                print("SOCKET LOST → RECONNECTING:", e)
                try:
                    conn.close()
                except:
                    pass
                self.msleep(2000)   # wait 2 sec and reconnect




    def resume(self):
        print("SOCKET RESUMED")
        self.waiting_for_user = False

    def stop(self):
        self.running = False


    def pause(self):
        print("SOCKET PAUSED")
        self.waiting_for_user = True




# ================= DB INIT =================
def init_db():
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()
    cur.execute(
        f"""
        CREATE TABLE IF NOT EXISTS {TABLE} (
            id SERIAL PRIMARY KEY,
            employee_id TEXT,
            work_order TEXT,
            charge_no TEXT,
            serial_no TEXT,
            part_no TEXT,
            unique_no TEXT,
            status TEXT,
            time TIMESTAMP,
            image BYTEA
        )
    """
    )
    conn.commit()
    cur.close()
    conn.close()


# ================= DB SAVE =================
def save_record(data, status, img_bytes):
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()
    cur.execute(
        f"""
        INSERT INTO {TABLE}
        (employee_id, work_order, charge_no, serial_no,
         part_no, unique_no, status, time, image)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
    """,
        (
            data["emp"],
            data["wo"],
            data["charge"],
            data["serial"],
            data["part"],
            data["unique"],
            status,
            datetime.now(),
            psycopg2.Binary(img_bytes),
        ),
    )
    conn.commit()
    cur.close()
    conn.close()


# ================= DB FETCH =================
def fetch_report(from_dt, to_dt, status):
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()
    q = f"""
        SELECT employee_id, work_order, charge_no,
               serial_no, part_no, unique_no,
               image, status, time
        FROM {TABLE}
        WHERE time BETWEEN %s AND %s
    """
    params = [from_dt, to_dt]
    if status != "ALL":
        q += " AND status=%s"
        params.append(status)
    q += " ORDER BY time DESC"
    cur.execute(q, params)
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return rows


def get_home_counts():
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    # Total counts
    cur.execute(
        f"""
        SELECT
            COUNT(*) AS total,
            COUNT(*) FILTER (WHERE status='OK') AS ok_count,
            COUNT(*) FILTER (WHERE status='NOT_OK') AS not_ok_count
        FROM {TABLE}
    """
    )
    total, ok_cnt, not_ok_cnt = cur.fetchone()

    # Today count
    today_start = datetime.combine(datetime.today().date(), time.min)
    today_end = datetime.combine(datetime.today().date(), time.max)

    cur.execute(
        f"""
        SELECT COUNT(*)
        FROM {TABLE}
        WHERE time BETWEEN %s AND %s
    """,
        (today_start, today_end),
    )
    today_cnt = cur.fetchone()[0]

    cur.close()
    conn.close()

    return (total or 0, ok_cnt or 0, not_ok_cnt or 0, today_cnt or 0)


# ================= CONFIRM DIALOG =================
class ConfirmDialog(QDialog):
    decision = Signal(str)

    def __init__(self, pix):
        super().__init__()
        self.setWindowTitle("Confirm Capture")

        img = QLabel(alignment=Qt.AlignCenter)
        img.setPixmap(pix.scaled(520, 360, Qt.KeepAspectRatio))

        ok = QPushButton("OK")
        nok = QPushButton("NOT OK")

        ok.setStyleSheet("background:#28a745;color:white;padding:12px;font-size:16px;")
        nok.setStyleSheet("background:#dc3545;color:white;padding:12px;font-size:16px;")

        ok.clicked.connect(lambda: self.finish("OK"))
        nok.clicked.connect(lambda: self.finish("NOT_OK"))

        lay = QVBoxLayout(self)
        lay.addWidget(img)
        btns = QHBoxLayout()
        btns.addWidget(ok)
        btns.addWidget(nok)
        lay.addLayout(btns)

    def finish(self, res):
        self.decision.emit(res)
        self.accept()


# ================= HOME =================
class Home(QWidget):
    def __init__(self):
        super().__init__()

        self.setStyleSheet(
            """
        QWidget {
            background: #f4f6f9;
            font-family: Segoe UI;
        }
        QLabel {
            color: #333;
        }
        """
        )

        self.main = QVBoxLayout(self)
        self.main.setContentsMargins(30, 30, 30, 30)
        self.main.setSpacing(20)

        title = QLabel("Home")
        title.setStyleSheet("font-size:22px;font-weight:bold;")
        self.main.addWidget(title)

        # ---- Cards ----
        self.cards = QHBoxLayout()
        self.cards.setSpacing(20)
        self.main.addLayout(self.cards)

        self.total_lbl = self._card("TOTAL INSPECTIONS", "#1e88e5")
        self.ok_lbl = self._card("OK COUNT", "#2e7d32")
        self.nok_lbl = self._card("NOT OK COUNT", "#c62828")
        self.today_lbl = self._card("TODAY INSPECTIONS", "#6a1b9a")

        self.main.addStretch()

        self.refresh()  # initial load

    def _card(self, title, color):
        w = QWidget()
        w.setStyleSheet("background:white;border-radius:8px;")
        v = QVBoxLayout(w)
        v.setContentsMargins(20, 20, 20, 20)

        t = QLabel(title)
        t.setStyleSheet("color:#777;font-size:13px;")

        val = QLabel("0")
        val.setStyleSheet(f"font-size:22px;font-weight:bold;color:{color};")

        v.addWidget(t)
        v.addWidget(val)

        self.cards.addWidget(w)
        return val

    def refresh(self):
        total, ok_cnt, not_ok_cnt, today_cnt = get_home_counts()
        self.total_lbl.setText(str(total))
        self.ok_lbl.setText(str(ok_cnt))
        self.nok_lbl.setText(str(not_ok_cnt))
        self.today_lbl.setText(str(today_cnt))

    def showEvent(self, event):
        super().showEvent(event)


# ================= OPERATOR =================
class Operator(QWidget):
    record_saved = Signal()
    EMP_LEN = 10
    WO_LEN = 10

    def __init__(self):
        super().__init__()

        # ---- Dark UI, clean inputs ----
        self.setStyleSheet(
            """
        QWidget { background:#121212; }
        QLabel { color:#ddd; background:transparent; }
        QLineEdit {
            background:transparent;
            color:black;
            border:1px solid #555;
            padding:6px;
            font-size:13px;
        }
        """
        )

        self.cap = None
        self.frame = None

        # ---- Top bar with refresh ----
        self.btn_refresh = QPushButton("🔄 New User")
        self.btn_refresh.setStyleSheet(
            "background:#6c757d;color:white;font-weight:bold;padding:6px 16px;"
        )
        self.btn_refresh.clicked.connect(self.reset_all)

        top = QHBoxLayout()
        top.addStretch()
        top.addWidget(self.btn_refresh)

        # ---- Employee & Work Order ----

        self.emp = QLineEdit(placeholderText="Employee ID")
        self.wo = QLineEdit(placeholderText="Work Order")

        normal = """
        QLineEdit {
            background:white;
            color:black;
            border:1px solid #999;
        }
        """
        self.emp.setStyleSheet(normal)
        self.wo.setStyleSheet(normal)

        self.emp.returnPressed.connect(self.emp_done)
        self.wo.returnPressed.connect(self.wo_done)
        self.emp.hide()
        self.wo.hide()


        # ---- 4 fields ----
        self.fields = {
            "charge": ("charge no", 14),
            "unique": ("unique no", 4),
            "Vendor Code": ("Vendor code", 8),
            "serial": ("serial no", 3),
        }

        self.inputs = {}
        grid = QGridLayout()
        grid.setHorizontalSpacing(12)
        grid.setVerticalSpacing(4)

        col = 0
        for key, (label, ln) in self.fields.items():
            lbl = QLabel(label)
            lbl.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
            lbl.setStyleSheet("font-size:11px;color:#ccc;")

            le = QLineEdit()

            le.setFixedWidth(160)
            le.setMaxLength(ln)
            le.setPlaceholderText(f"{ln} digits")
            le.textChanged.connect(lambda _, k=key: self.validate_field(k))

            lbl.hide()
            le.hide()

            grid.addWidget(lbl, 0, col)
            grid.addWidget(le, 1, col)
            self.inputs[key] = (lbl, le, ln)
            col += 1

        # ---- Camera ----
        self.preview = QLabel("Camera OFF")
        self.preview.setAlignment(Qt.AlignCenter)
        self.preview.setMinimumHeight(420)
        self.preview.setStyleSheet("background:black;color:white;")

        self.timer = QTimer(self)
        self.timer.timeout.connect(self.update_frame)

        lay = QVBoxLayout(self)
        lay.addLayout(top)
        lay.addWidget(self.emp)
        lay.addWidget(self.wo)
        lay.addLayout(grid)
        lay.addWidget(self.preview)

        def all_fields_valid(self):
            for _, le, ln in self.inputs.values():
                if len(le.text()) != ln:
                    return False
            return True


    def try_capture(self):
        print("ENTER PRESSED")

        if self.frame is None:
            print("❌ Camera frame not ready")
            return

        if not self.all_fields_valid():
            print("❌ Fields invalid")
            return

        print("✅ CAPTURE TRIGGERED")
        self.capture()





    def on_socket_data(self, msg):
        print("Socket accepted:", msg)

        charge = msg[0:14]
        unique = msg[14:21]
        serial = msg[-5:-2]
        vendor_code = "16099680"

        self.inputs["charge"][1].setText(charge)
        self.inputs["unique"][1].setText(unique)
        self.inputs["serial"][1].setText(serial)
        self.inputs["Vendor Code"][1].setText(vendor_code)

        #  VERY IMPORTANT
        self.inputs["serial"][1].setFocus()









    # ---------- FLOW ----------
    def emp_done(self):
        if self.emp.text():
            self.emp.hide()
            self.wo.show()
            self.wo.setFocus()

    def wo_done(self):
        if self.wo.text():
            self.wo.hide()

            for lbl, le, _ in self.inputs.values():
                lbl.show()
                le.show()

            self.start_camera()

            if not hasattr(self, "socket_thread"):
                self.socket_thread = FHVSocketThread()
                self.socket_thread.data_received.connect(self.on_socket_data)
                self.socket_thread.start()



    def validate_field(self, key):
        _, le, ln = self.inputs[key]
        if not le.text():
            le.setStyleSheet(
                "background:transparent;color:white;border:1px solid #555;"
            )
        elif len(le.text()) == ln:
            le.setStyleSheet("background:#1b5e20;color:white;border:1px solid #2ecc71;")
        else:
            le.setStyleSheet("background:#7f0000;color:white;border:1px solid #e74c3c;")

    # ---------- CAMERA ----------
    def start_camera(self):
        if self.cap is None:
            self.cap = cv2.VideoCapture(1, cv2.CAP_DSHOW)
            if not self.cap.isOpened():
                self.preview.setText("Camera not available")
                return
        self.timer.start(20)

    def stop_camera(self):
        self.timer.stop()
        if self.cap:
            self.cap.release()
            self.cap = None
        self.preview.setText("Camera OFF")

    def update_frame(self):
        if not self.cap:
            return

        ret, frame = self.cap.read()
        if not ret:
            return

        self.frame = frame
        rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        h, w, ch = rgb.shape
        self.preview.setPixmap(
            QPixmap.fromImage(QImage(rgb.data, w, h, ch * w, QImage.Format_RGB888))
        )

    # ---------- RESET ----------
    def reset_all(self):
        self.stop_camera()

        self.emp.clear()
        self.wo.clear()

        self.emp.show()
        self.wo.hide()
        self.socket_waiting = False


        for lbl, le, _ in self.inputs.values():
            lbl.hide()
            le.hide()
            le.clear()
            le.setStyleSheet(
                "background:transparent;color:white;border:1px solid #555;"
            )

        self.emp.setFocus()

    def all_fields_valid(self):
        for _, le, ln in self.inputs.values():
            if len(le.text()) != ln:
                return False
        return True




    def on_enter(self):
        # Called when Operator page is shown
        if hasattr(self, "socket_thread"):
            print("Operator screen → resume socket")
            self.socket_thread.resume()

    def on_leave(self):
        # Called when Operator page is hidden
        if hasattr(self, "socket_thread"):
            print("Leaving Operator → pause socket")
            self.socket_thread.pause()





    # ---------- CAPTURE ----------
    def keyPressEvent(self, e):
        if e.key() == Qt.Key_Return and self.frame is not None:
            if not self.all_fields_valid():
                return
            self.capture()









    def capture(self):
            _, buf = cv2.imencode(".jpg", self.frame)
            img_bytes = buf.tobytes()

            #  CLEAN *EVERY* FIELD BEFORE DB
            data = {
                "emp": clean_text(self.emp.text()),
                "wo": clean_text(self.wo.text()),
                "charge": clean_text(self.inputs["charge"][1].text()),
                "serial": clean_text(self.inputs["serial"][1].text()),
                "part": clean_text(self.inputs["Vendor Code"][1].text()),
                "unique": clean_text(self.inputs["unique"][1].text()),
            }

            pix = QPixmap.fromImage(QImage.fromData(img_bytes))
            dlg = ConfirmDialog(pix)

            def after_save(res):
                save_record(data, res, img_bytes)

                for _, le, _ in self.inputs.values():
                    le.clear()

                if hasattr(self, "socket_thread"):
                    self.socket_thread.resume()

                self.record_saved.emit()

            dlg.decision.connect(after_save)
            dlg.exec()





# ================= REPORT =================
class Report(QWidget):
    def __init__(self):
        super().__init__()
        main = QVBoxLayout(self)

        header = QHBoxLayout()

        left = QHBoxLayout()
        self.from_dt = QDateEdit(calendarPopup=True)
        self.from_dt.setDate(QDate.currentDate().addDays(-7))
        self.to_dt = QDateEdit(calendarPopup=True)
        self.to_dt.setDate(QDate.currentDate())
        self.status = QComboBox()
        self.status.addItems(["ALL", "OK", "NOT_OK"])

        for w in ("From", self.from_dt, "To", self.to_dt, "Status", self.status):
            left.addWidget(QLabel(w) if isinstance(w, str) else w)

        right = QHBoxLayout()
        self.btn_excel = QPushButton("Export Excel")
        self.btn_excel.setStyleSheet(
            "background:#28a745;color:white;font-weight:bold;padding:6px 18px;"
        )
        right.addStretch()
        right.addWidget(self.btn_excel)

        header.addLayout(left)
        header.addStretch()
        header.addLayout(right)
        main.addLayout(header)

        self.table = QTableWidget(0, 10)
        self.table.setHorizontalHeaderLabels(
            [
                "Employee ID",
                "Work Order",
                "Charge No",
                "Serial No",
                "Vendor Code",
                "Batch No",
                "Image",
                "Status",
                "Date",
                "Time",
            ]
        )
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.verticalHeader().setVisible(False)
        main.addWidget(self.table)

        self.from_dt.dateChanged.connect(self.load)
        self.to_dt.dateChanged.connect(self.load)
        self.status.currentIndexChanged.connect(self.load)
        self.btn_excel.clicked.connect(self.export_excel)

        self.load()

    def load(self):
        self.table.setRowCount(0)  # clear table first

        f = datetime.combine(self.from_dt.date().toPython(), time.min)
        t = datetime.combine(self.to_dt.date().toPython(), time.max)
        rows = fetch_report(f, t, self.status.currentText())

        for r in rows:
            row = self.table.rowCount()
            self.table.insertRow(row)
            self.table.setRowHeight(row, 160)

            # Only add text first
            for c in range(6):
                item = QTableWidgetItem(str(r[c]))
                item.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(row, c, item)

            st = QTableWidgetItem(r[7])
            st.setTextAlignment(Qt.AlignCenter)
            st.setForeground(QColor("green") if r[7] == "OK" else QColor("red"))
            self.table.setItem(row, 7, st)

            d = QTableWidgetItem(r[8].strftime("%Y-%m-%d"))
            t = QTableWidgetItem(r[8].strftime("%H:%M:%S"))

            d.setTextAlignment(Qt.AlignCenter)
            t.setTextAlignment(Qt.AlignCenter)

            self.table.setItem(row, 8, d)
            self.table.setItem(row, 9, t)

            self.table.horizontalHeader().setStyleSheet(
                """
                                                                QHeaderView::section {
                                                                    background-color: #1e88e5;
                                                                    color: white;
                                                                    font-weight: bold;
                                                                    padding: 6px;
                                                                    border: none;
                                                                }
                                                                """
            )

            # Load image asynchronously after small delay
            QTimer.singleShot(10, lambda row=row, img=r[6]: self._set_image(row, img))

    def _set_image(self, row, img_bytes):
        pix = QPixmap()
        pix.loadFromData(bytes(img_bytes))
        lbl = QLabel(alignment=Qt.AlignCenter)
        lbl.setPixmap(pix.scaled(240, 140, Qt.KeepAspectRatio))
        self.table.setCellWidget(row, 6, lbl)

    def export_excel(self):
        path, _ = QFileDialog.getSaveFileName(self, "Save Excel", "", "Excel (*.xlsx)")
        if not path:
            return

        wb = Workbook()
        ws = wb.active
        ws.append(
            [
                "Employee ID",
                "Work Order",
                "Charge No",
                "Serial No",
                "Part No",
                "Unique No",
                "Status",
                "Date",
                "Time",
            ]
        )

        green = PatternFill("solid", fgColor="C6EFCE")
        red = PatternFill("solid", fgColor="FFC7CE")

        rows = fetch_report(
            datetime.combine(self.from_dt.date().toPython(), time.min),
            datetime.combine(self.to_dt.date().toPython(), time.max),
            self.status.currentText(),
        )

        for r in rows:
            ws.append(
                [
                    r[0],
                    r[1],
                    r[2],
                    r[3],
                    r[4],
                    r[5],
                    r[7],
                    r[8].strftime("%Y-%m-%d"),
                    r[8].strftime("%H:%M:%S"),
                ]
            )
            ws[f"G{ws.max_row}"].fill = green if r[7] == "OK" else red

        wb.save(path)


# ================= MAIN =================
class Main(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Camera Inspection System")
        self.resize(1600, 900)

        nav = QHBoxLayout()
        self.stack = QStackedWidget()

        # ---- Pages ----
        self.home = Home()
        self.operator = Operator()
        self.report = Report()

        # ---- Add pages to stack (CRITICAL) ----
        self.stack.addWidget(self.home)
        self.stack.addWidget(self.operator)
        self.stack.addWidget(self.report)

        # ---- Navigation buttons (NO UI change) ----
        btn_home = QPushButton("Home")
        btn_operator = QPushButton("Operator")
        btn_report = QPushButton("Report")

        btn_home.clicked.connect(self.go_home)
        btn_operator.clicked.connect(self.go_operator)
        btn_report.clicked.connect(self.go_report)


        nav.addWidget(btn_home)
        nav.addWidget(btn_operator)
        nav.addWidget(btn_report)
        nav.addStretch()

        # ---- Refresh connections ----
        self.operator.record_saved.connect(self.report.load)
        self.operator.record_saved.connect(self.home.refresh)

        # ---- Layout ----
        lay = QVBoxLayout(self)
        lay.addLayout(nav)
        lay.addWidget(self.stack)

        # ---- Default page ----
        self.stack.setCurrentWidget(self.home)



    def go_home(self):
            self.operator.on_leave()
            self.stack.setCurrentWidget(self.home)

    def go_operator(self):
            self.stack.setCurrentWidget(self.operator)
            self.operator.on_enter()

    def go_report(self):
            self.operator.on_leave()
            self.stack.setCurrentWidget(self.report)




# ================= RUN =================
if __name__ == "__main__":
    init_db()
    app = QApplication(sys.argv)
    w = Main()
    w.show()
    sys.exit(app.exec())
